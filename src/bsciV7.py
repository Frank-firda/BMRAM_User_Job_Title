import win32com.client
import ctypes
import winsound
import os
from datetime import datetime
from tqdm import tqdm
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time
import json
import sys
import psutil


app_data_folder = os.getenv('LOCALAPPDATA')
file_path = os.path.join(app_data_folder, "BMRAM_User_Job_Title\\BMRAM User, JobTitle_data\\settings.json")



if os.path.exists(file_path):
    with open(file_path, 'r') as json_file:
        settings = json.load(json_file)
else:
    data = {    
    "read_col": "A",
    "write_Email": "B",
    "write_Title": "C",
    "should_write_email": True,
    "should_write_title": True,
    "add_drop_down_filter": True,
    "open_when_done": False,
    "halt_on_error": False,
    "create_new_excel":False,
    "sheet": "Sheet1"
    }
    settings = data
    with open(file_path, 'w') as json_file:
        json.dump(data, json_file, indent=4)

print(settings)

def column_string_to_index(column_string):
    column_string = column_string.upper()
    index = 0
    for i, char in enumerate(reversed(column_string)):
        index += (ord(char) - ord('A') + 1) * (26 ** i)
    return index - 1

def show_windows_message(title, message):
    """Display a Windows message box with an OK button."""
    winsound.MessageBeep()
    ctypes.windll.user32.MessageBoxW(0, message, title, 0)

iread_col = settings["read_col"]
iwrite_Email = settings["write_Email"]
iwrite_Title = settings["write_Title"]
ishould_write_email = settings["should_write_email"]
ishould_write_title = settings["should_write_title"]
iadd_drop_down_filter = settings["add_drop_down_filter"]
iopen_when_done = settings["open_when_done"]
ihalt_on_error = settings["halt_on_error"]
icreate_new_excel = settings["create_new_excel"]
isheet = settings["sheet"]




def mainscript(file_path):
    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")
    address_list = namespace.AddressLists.Item("Global Address List")  # Ensure this name matches
    gal = address_list.AddressEntries
    start_time = datetime.now()

    df = pd.read_excel(file_path, sheet_name=isheet)  
    wb = load_workbook(file_path)






    ws = wb[isheet]

    if ishould_write_title:
        ws[iwrite_Title+"1"] = "Job Title"
    if ishould_write_email:
        ws[iwrite_Email+"1"] = "Email"

    if iadd_drop_down_filter:
        sheet = wb.active
        sheet.auto_filter.ref = sheet.dimensions

    

    names_data = []
    for index, value in df.iloc[:, column_string_to_index(iread_col)].items():  
        if pd.notna(value):  
        
            try:
                last_name, first_name = value.split(", ")
                names_data.append({
                    'Last Name': last_name.strip(),
                    'First Name': first_name.strip(),
                    'Position': index + 2  
                })
            except ValueError:
                print(f"Skipping invalid format in row {index + 2}: {value}")

    names_df = pd.DataFrame(names_data)
    for index, row in names_df.iterrows():
        last_name = row['Last Name']
        first_name = row['First Name']
        position = row['Position']
    
    longestName = 0
    longestTitle = 0
    longestEmail = 0

    def get_position(first_name, last_name):
        found = names_df[(names_df['First Name'].str.lower() == first_name.lower()) &
                     (names_df['Last Name'].str.lower() == last_name.lower())]
    
        if not found.empty:
            return found['Position'].values[0]
        else:
            return None
        
        
              


    total_items = len(names_data)
    pbar = tqdm(names_data, desc="Processing", unit=" contact", ncols=100)

    for entry in pbar:
        try:
            first_name = entry['First Name']
            last_name = entry['Last Name']
            fullName = last_name+", "+first_name
            contact = gal.Item(fullName)
            exchange_user = contact.GetExchangeUser()
            if exchange_user is None:
               
                print(fullName+" was not fould in the GAL!")
            con_pos = get_position(first_name, last_name)
            shouldPass = False
            dispName = getattr(exchange_user,"Name", None)
            if first_name.lower() in dispName.lower() and last_name.lower() in dispName.lower():
                pass
            else:
                if ihalt_on_error:
                    show_windows_message("Information", fullName+" Not found")
                    if ishould_write_title:
                        ws[iwrite_Title+str(con_pos)] = "Not Found"
                        
                    if ishould_write_email:
                        ws[iwrite_Email+str(con_pos)] = "Not Found"
                        
                    continue
                else:
                    print("No match found for: "+fullName)
                    if ishould_write_title:
                        ws[iwrite_Title+str(con_pos)] = "Not Found"
                        

                    if ishould_write_email:
                        ws[iwrite_Email+str(con_pos)] = "Not Found"
                           
                    continue



            if ishould_write_email:

                primary_smtp = getattr(exchange_user, "PrimarySmtpAddress", None)

                while primary_smtp is None:
                    primary_smtp = getattr(exchange_user, "PrimarySmtpAddress", None)
                    print("cant find primary_smtp for: "+fullName+"\n")
                    show_windows_message("Information", "User Action Required")
                    option = input("R: retry"+"\n"+"S: skip"+"\n"+"C: cancel processe"+"\n"+"enter: ")
                    
                    if option.strip().lower() == "r":
                        continue
                    elif option.strip().lower() == "s":
                        shouldPass = True
                        break
                    elif option.strip().lower() == "c":
                        sys.exit(0)

                    
            if shouldPass == True:
                continue


            if ishould_write_title:
                job_title = getattr(exchange_user, "JobTitle", None)
                while job_title is None:
                    job_title = getattr(exchange_user, "JobTitle", None)
                    print("cant find job title for: "+fullName+"\n")
                    show_windows_message("Information", "User Action Required")
                    option = input("R: retry"+"\n"+"S: skip"+"\n"+"C: cancel processe"+"\n"+"enter: ")

                    if option.strip().lower() == "r":
                        continue
                    elif option.strip().lower() == "s":
                        shouldPass = True
                        break
                    elif option.strip().lower() == "c":
                        sys.exit(0)

            if shouldPass == True:
                continue

            
            if ishould_write_title:
                ws[iwrite_Title+str(con_pos)] = job_title

            if ishould_write_email:
                ws[iwrite_Email+str(con_pos)] = primary_smtp

            if len(fullName) > longestName:
                longestName = len(fullName)
            if ishould_write_email:
                if primary_smtp and len(primary_smtp) > longestEmail:
                    longestEmail = len(primary_smtp)
            if ishould_write_title:
                if job_title and len(job_title) > longestTitle:
                    longestTitle = len(job_title)
        except Exception as e:
            print(e)
    

    ws.column_dimensions[iread_col].width = longestName
    if ishould_write_email:
        ws.column_dimensions[iwrite_Email].width = longestEmail
    if ishould_write_title:
        ws.column_dimensions[iwrite_Title].width = longestTitle

    if icreate_new_excel:
        file_path = filedialog.asksaveasfilename(
        title="Save a file",
        defaultextension=".xlsx",  
        filetypes=[("Excel files", "*.xlsx")],  
        )
    
        if file_path:
            wb.save(file_path)
            print("time taken: ", datetime.now() - start_time)
            show_windows_message("Information", "process completed successfully!")
    else:
        wb.save(file_path)
        print("time taken: ", datetime.now() - start_time)
        show_windows_message("Information", "process completed successfully!")
    
    if iopen_when_done:
        os.startfile(file_path)

def select_file():
    root = tk.Tk()
    root.withdraw()  
   
    file_path = filedialog.askopenfilename(title="Select an Excel file",
        filetypes=[("Excel files", "*.xlsx")])

    if file_path:
        print(f"Selected file: {file_path}")
        mainscript(file_path)

    else:
        print("No file selected.")

if __name__ == "__main__":
    select_file()